"""One-time import: legacy U-15 xlsx -> current DB taxonomy. Run: python scripts/import_u15_legacy.py [--dry-run]"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from backend import repository as repo
from backend.csv_import import _build_lookup_context, _is_placeholder
from backend.db import DEFAULT_DB_PATH
from backend.format_time import parse_timestamp
from backend.match_score import calculate_match_score_for_match

XLSX_PATH = PROJECT_ROOT / "Rugby U-15 stats (1).xlsx"
MATCH_MAP = {1: 6, 2: 5, 3: 7, 4: 8, 5: 9}

LEGACY_TO_CURRENT: dict[tuple[str, str], tuple[str, str, str | None]] = {
    ("Handling", "Pass (success)"): ("Handling", "Pass", "Success"),
    ("Handling", "Pass (missed)"): ("Handling", "Pass", "Failure"),
    ("Handling", "Catch (success)"): ("Handling", "Catch", "Success"),
    ("Handling", "Catch (failure)"): ("Handling", "Catch", "Failure"),
    ("Handling", "Knock-on"): ("Handling", "Knock-on", None),
    ("Handling", "Forward pass"): ("Handling", "Forward pass", None),
    ("Offence", "Carry"): ("Offence", "Carry", None),
    ("Offence", "Iinebreak"): ("Offence", "Linebreak", None),
    ("Offence", "try"): ("Offence", "Try", None),
    ("Tackle", "Tackle (success)"): ("Tackle", "Tackle", "Success"),
    ("Tackle", "Missed tackle"): ("Tackle", "Tackle", "Failure"),
    ("Kicking", "Kick (tactical)"): ("Kicking", "Kick (tactical)", None),
    ("Kicking", "Kick (attacking)"): ("Kicking", "Kick (attacking)", None),
    ("Kicking", "conversion"): ("Kicking", "Conversion", None),
    ("Kicking", "Conversion (opp)"): ("Kicking", "Conversion", None),
    ("Set-piece", "Lineout (own)"): ("Set-piece", "Lineout (own)", None),
    ("Set-piece", "Lineout (opp)"): ("Set-piece", "Lineout (opp)", None),
    ("Set-piece", "Scrum (own)"): ("Set-piece", "Scrum (own)", None),
    ("Set-piece", "Scrum (opp)"): ("Set-piece", "Scrum (opp)", None),
    ("Set-piece", "Ruck"): ("Team play", "Ruck", None),
    ("Defence", "Turnover"): ("Defence", "Turnover", None),
    ("Defence", "Try conceded"): ("Defence", "Turnover", None),
    ("Discipline", "penalty (offside)"): ("Discipline", "Penalty", None),
    ("Discipline", "penalty (high tackle)"): ("Discipline", "Penalty", None),
    ("Discipline", "penalty (ruck)"): ("Discipline", "Penalty", None),
    ("Discipline", "penalty (other)"): ("Discipline", "Penalty", None),
    ("Discipline", "Penalty (opponent)"): ("Discipline", "Penalty", None),
    ("Discipline", "yellow card"): ("Discipline", "yellow card", None),
    ("Discipline", "red card"): ("Discipline", "red card", None),
    ("Substitute", "On"): ("Substitute", "On", None),
    ("Substitute", "Off"): ("Substitute", "Off", None),
}


def _norm(text: str) -> str:
    return text.strip().lower()


def _optional_outcome(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if _is_placeholder(text):
        return None
    return text


def load_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        row = {headers[i]: ("" if raw[i] is None else str(raw[i]).strip()) for i in range(len(headers))}
        rows.append(row)
    wb.close()
    return rows


def map_legacy_row(cat: str, ev: str, outcome: str) -> tuple[str, str, str | None] | None:
    mapped = LEGACY_TO_CURRENT.get((cat.strip(), ev.strip()))
    if not mapped:
        return None
    new_cat, new_act, outcome_override = mapped
    if outcome_override is not None:
        return new_cat, new_act, outcome_override
    return new_cat, new_act, _optional_outcome(outcome)


def normalize_subject(player: str, cat: str, ev: str) -> str:
    if _is_placeholder(player):
        if cat == "Set-piece" or ev == "Ruck":
            return "Команда"
    return player


def prepare_rows(
    conn: sqlite3.Connection,
    rows: list[dict[str, str]],
) -> tuple[list[dict], list[str]]:
    prepared: list[dict] = []
    errors: list[str] = []
    contexts: dict[int, tuple] = {}

    for db_mid in set(MATCH_MAP.values()):
        ctx = _build_lookup_context(conn, db_mid)
        if ctx:
            contexts[db_mid] = ctx
        else:
            errors.append(f"Матч {db_mid} не найден в БД")

    for idx, row in enumerate(rows, start=2):
        file_mid = int((row.get("Match_ID") or "0").strip())
        db_mid = MATCH_MAP.get(file_mid)
        if not db_mid or db_mid not in contexts:
            if file_mid not in MATCH_MAP:
                errors.append(f"Строка {idx}: неизвестный Match_ID={file_mid}")
            continue

        cat = (row.get("Event_Category") or "").strip()
        ev = (row.get("Event") or "").strip()
        player = normalize_subject((row.get("Player_ID") or "").strip(), cat, ev)
        outcome = (row.get("Outcome") or "").strip()

        if _is_placeholder(cat) and _is_placeholder(ev):
            continue

        mapped = map_legacy_row(cat, ev, outcome)
        if not mapped:
            errors.append(f"Строка {idx} M{file_mid}: нет маппинга {cat}/{ev}")
            continue
        new_cat, new_act, new_outcome = mapped

        _, action_map, player_lookup, team_lookup = contexts[db_mid]
        try:
            period = int((row.get("Half") or "").strip())
            time_sec = parse_timestamp((row.get("Time_stamp") or "").strip())
        except ValueError:
            errors.append(f"Строка {idx}: некорректный тайм/время")
            continue

        action = action_map.get((_norm(new_cat), _norm(new_act)))
        if not action:
            errors.append(f"Строка {idx} M{file_mid}: в шаблоне нет {new_cat}/{new_act}")
            continue

        player_id = None
        team_id = None
        subject_type = "player"
        if player and not _is_placeholder(player):
            pl = player_lookup.get(_norm(player))
            tm = team_lookup.get(_norm(player))
            if not pl and not tm:
                errors.append(f"Строка {idx} M{file_mid}: субъект '{player}' не найден")
                continue
            subject_type = "player" if pl else "team"
            player_id = int(pl["Id"]) if pl else None
            team_id = int(tm["Id"]) if tm else None
        else:
            errors.append(f"Строка {idx} M{file_mid} {cat}/{ev}: не задан субъект")
            continue

        prepared.append(
            {
                "line": idx,
                "match_id": db_mid,
                "period": period,
                "time": time_sec,
                "subject_type": subject_type,
                "player_id": player_id,
                "team_id": team_id,
                "action_id": int(action["Id"]),
                "outcome": new_outcome,
                "comment": None,
            }
        )

    return prepared, errors


def import_events(conn: sqlite3.Connection, prepared: list[dict]) -> int:
    with conn:
        for item in prepared:
            conn.execute(
                """INSERT INTO Event (MatchId, PeriodNumber, TimestampSec, SubjectType, PlayerId, TeamId, ActionId, Outcome, Comment, CreatedAt)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))""",
                (
                    item["match_id"],
                    item["period"],
                    item["time"],
                    item["subject_type"],
                    item["player_id"],
                    item["team_id"],
                    item["action_id"],
                    item["outcome"],
                    item["comment"],
                ),
            )
    return len(prepared)


def print_summary(conn: sqlite3.Connection, prepared: list[dict]) -> None:
    inv = {v: k for k, v in MATCH_MAP.items()}
    by_match: dict[int, int] = {}
    for item in prepared:
        by_match[item["match_id"]] = by_match.get(item["match_id"], 0) + 1

    print("\n--- Итог по матчам ---")
    for db_mid in sorted(MATCH_MAP.values()):
        m = repo.get_match(conn, db_mid)
        ht = repo.get_team(conn, int(m["HomeTeamId"]))
        at = repo.get_team(conn, int(m["AwayTeamId"]))
        sc = calculate_match_score_for_match(conn, db_mid)
        print(
            f"  DB {db_mid} (file M{inv[db_mid]}): {ht['Name']} {sc['home']} : {sc['away']} {at['Name']}"
            f"  |  +{by_match.get(db_mid, 0)} событий"
        )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    args = parser.parse_args()

    if not XLSX_PATH.is_file():
        print(f"Файл не найден: {XLSX_PATH}", file=sys.stderr)
        return 1

    rows = load_rows(XLSX_PATH)
    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row

    for db_mid in MATCH_MAP.values():
        existing = repo.count_events_by_match(conn, db_mid)
        if existing and not args.dry_run:
            print(f"Матч {db_mid} уже содержит {existing} событий. Прервите или удалите вручную.", file=sys.stderr)
            return 1

    prepared, errors = prepare_rows(conn, rows)
    print(f"Строк в файле: {len(rows)}")
    print(f"Готово к импорту: {len(prepared)}")
    print(f"Ошибок: {len(errors)}")
    if errors:
        for err in errors:
            print(f"  {err}")
        return 1

    if args.dry_run:
        print("\n[dry-run] Запись в БД не выполнялась.")
        print_summary(conn, prepared)
        return 0

    count = import_events(conn, prepared)
    print(f"\nИмпортировано: {count} событий")
    print_summary(conn, prepared)
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
